from openpyxl import load_workbook
import glob
import os
import easygui


table_dir = "D:\\jakub.kaleta\\automat\\iveco\\excel\\"
# myfile = table_dir + "tabela_iveco_matryca.xlsx"
myfile = easygui.fileopenbox()
workbook_baza = load_workbook(filename=myfile, data_only=True)
sheet_baza = workbook_baza.active

co1 = 0.7
co2 = 0.2
co3 = 0.1


def safe_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def compare(cur_sheet):
    scr_show = 0.0
    scr_ang = 0.0
    scr_pianka = 0.0

    for i in range(2, 82):
        cur_x = safe_float(cur_sheet["A" + str(i)].value)
        baza_x = safe_float(sheet_baza["A" + str(i)].value)
        cur_y = safe_float(cur_sheet["B" + str(i)].value)
        baza_y = safe_float(sheet_baza["B" + str(i)].value)
        cur_z = safe_float(cur_sheet["C" + str(i)].value)
        baza_z = safe_float(sheet_baza["C" + str(i)].value)
        if abs(cur_x - baza_x) < 1 and abs(cur_y - baza_y) < 1 and abs(cur_z - baza_z) < 1 and \
                cur_sheet["D" + str(i)].value == sheet_baza["D" + str(i)].value:
            scr_show += 1.0

    for i in range(2, 82):
        if cur_sheet["E" + str(i)].value == sheet_baza["E" + str(i)].value:
            scr_show += 1.0

    for i in range(2, 82):
        if cur_sheet["K" + str(i)].value == sheet_baza["K" + str(i)].value:
            scr_show += 1.0

    scr = co1*scr_show + co2*scr_ang + co3*scr_pianka
    return scr


files = glob.glob("D:\\jakub.kaleta\\automat\\iveco\\excel\\tabele\\*.xlsx")

res = {}
for file in files:
    basename = os.path.basename(file)
    workbook = load_workbook(filename=file, data_only=True)
    sheet = workbook.active
    score = round(compare(sheet), 2)
    res[basename] = score

res_sort = {k: v for k, v in sorted(res.items(), key=lambda item: item[1], reverse=True)}
for key, value in res_sort.items():
    print(key, ' : ', value)  # sort dict






